"""
VideoBot Pro - Export Utilities
Утилиты для экспорта данных в различные форматы
"""

import io
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
from fastapi.responses import StreamingResponse
import structlog

logger = structlog.get_logger(__name__)

async def export_users_to_csv(users: List[Any]) -> StreamingResponse:
    """Экспорт пользователей в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "Telegram ID", "Username", "First Name", "Last Name",
        "User Type", "Is Premium", "Is Banned", "Downloads Total",
        "Created At", "Last Active", "Premium Expires", "Language"
    ]
    writer.writerow(headers)
    
    # Данные
    for user in users:
        row = [
            user.id,
            user.telegram_id,
            user.username or "",
            user.first_name or "",
            user.last_name or "",
            user.user_type,
            "Yes" if user.is_premium else "No",
            "Yes" if user.is_banned else "No",
            user.downloads_total or 0,
            user.created_at.isoformat() if user.created_at else "",
            user.last_active_at.isoformat() if user.last_active_at else "",
            user.premium_expires_at.isoformat() if user.premium_expires_at else "",
            user.language_code or ""
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def export_users_to_excel(users: List[Any]) -> StreamingResponse:
    """Экспорт пользователей в Excel"""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl import Workbook
        
        # Подготавливаем данные
        data = []
        for user in users:
            data.append({
                "ID": user.id,
                "Telegram ID": user.telegram_id,
                "Username": user.username or "",
                "First Name": user.first_name or "",
                "Last Name": user.last_name or "",
                "Display Name": user.display_name,
                "User Type": user.user_type,
                "Is Premium": "Yes" if user.is_premium else "No",
                "Is Banned": "Yes" if user.is_banned else "No",
                "Downloads Total": user.downloads_total or 0,
                "Downloads Today": user.downloads_today or 0,
                "Created At": user.created_at,
                "Last Active": user.last_active_at,
                "Premium Expires": user.premium_expires_at,
                "Language": user.language_code or "",
                "Country": user.country_code or "",
                "Registration Source": user.registration_source or ""
            })
        
        # Создаём DataFrame
        df = pd.DataFrame(data)
        
        # Создаём Excel файл с форматированием
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Добавляем данные
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Форматирование заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except ImportError:
        # Если pandas/openpyxl не установлены, используем CSV
        logger.warning("pandas/openpyxl not available, falling back to CSV")
        return await export_users_to_csv(users)

async def export_downloads_to_csv(downloads: List[Any]) -> StreamingResponse:
    """Экспорт скачиваний в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "User ID", "Username", "Original URL", "Platform", 
        "Video Title", "Status", "Quality", "Format", "File Size MB",
        "Created At", "Started At", "Completed At", "Error Message"
    ]
    writer.writerow(headers)
    
    # Данные
    for download in downloads:
        file_size_mb = ""
        if hasattr(download, 'file_size_bytes') and download.file_size_bytes:
            file_size_mb = round(download.file_size_bytes / (1024 * 1024), 2)
        
        row = [
            download.id,
            download.user_id,
            getattr(download, 'username', ''),
            download.original_url,
            download.platform or "",
            download.video_title or "",
            download.status,
            download.actual_quality or "",
            download.actual_format or "",
            file_size_mb,
            download.created_at.isoformat() if download.created_at else "",
            download.started_at.isoformat() if download.started_at else "",
            download.completed_at.isoformat() if download.completed_at else "",
            download.error_message or ""
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=downloads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def export_downloads_to_excel(downloads: List[Any]) -> StreamingResponse:
    """Экспорт скачиваний в Excel"""
    try:
        import pandas as pd
        
        # Подготавливаем данные
        data = []
        for download in downloads:
            file_size_mb = None
            if hasattr(download, 'file_size_bytes') and download.file_size_bytes:
                file_size_mb = round(download.file_size_bytes / (1024 * 1024), 2)
            
            data.append({
                "ID": download.id,
                "User ID": download.user_id,
                "Username": getattr(download, 'username', ''),
                "User Type": getattr(download, 'user_type', ''),
                "Original URL": download.original_url,
                "Platform": download.platform or "",
                "Video Title": download.video_title or "",
                "Status": download.status,
                "Quality": download.actual_quality or "",
                "Format": download.actual_format or "",
                "File Size MB": file_size_mb,
                "Progress %": download.progress_percent or 0,
                "Retry Count": download.retry_count or 0,
                "Created At": download.created_at,
                "Started At": download.started_at,
                "Completed At": download.completed_at,
                "Processing Time": self._calculate_processing_time(download),
                "Error Message": download.error_message or ""
            })
        
        # Создаём DataFrame и Excel файл
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Downloads', index=False)
            
            # Форматирование
            workbook = writer.book
            worksheet = writer.sheets['Downloads']
            
            # Заморозка заголовков
            worksheet.freeze_panes = 'A2'
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=downloads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except ImportError:
        logger.warning("pandas not available, falling back to CSV")
        return await export_downloads_to_csv(downloads)

async def export_analytics_to_csv(analytics_data: Dict[str, Any]) -> StreamingResponse:
    """Экспорт аналитики в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Экспортируем различные секции аналитики
    
    # Общая статистика
    if 'overview' in analytics_data:
        writer.writerow(["=== ОБЩАЯ СТАТИСТИКА ==="])
        for key, value in analytics_data['overview'].items():
            writer.writerow([key, value])
        writer.writerow([])
    
    # Статистика пользователей
    if 'user_stats' in analytics_data:
        writer.writerow(["=== ПОЛЬЗОВАТЕЛИ ==="])
        for key, value in analytics_data['user_stats'].items():
            writer.writerow([key, value])
        writer.writerow([])
    
    # Статистика скачиваний
    if 'download_stats' in analytics_data:
        writer.writerow(["=== СКАЧИВАНИЯ ==="])
        for key, value in analytics_data['download_stats'].items():
            writer.writerow([key, value])
        writer.writerow([])
    
    # Дневная динамика
    if 'daily_data' in analytics_data:
        writer.writerow(["=== ДНЕВНАЯ ДИНАМИКА ==="])
        writer.writerow(["Дата", "Пользователи", "Скачивания", "Доходы"])
        for day_data in analytics_data['daily_data']:
            writer.writerow([
                day_data.get('date', ''),
                day_data.get('users', 0),
                day_data.get('downloads', 0),
                day_data.get('revenue', 0)
            ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def export_analytics_to_excel(analytics_data: Dict[str, Any]) -> StreamingResponse:
    """Экспорт аналитики в Excel с множественными листами"""
    try:
        import pandas as pd
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Лист с общей статистикой
            if 'overview' in analytics_data:
                overview_df = pd.DataFrame([
                    {"Метрика": k, "Значение": v} 
                    for k, v in analytics_data['overview'].items()
                ])
                overview_df.to_excel(writer, sheet_name='Overview', index=False)
            
            # Лист с пользователями
            if 'user_stats' in analytics_data:
                user_df = pd.DataFrame([
                    {"Метрика": k, "Значение": v} 
                    for k, v in analytics_data['user_stats'].items()
                ])
                user_df.to_excel(writer, sheet_name='Users', index=False)
            
            # Лист со скачиваниями
            if 'download_stats' in analytics_data:
                download_df = pd.DataFrame([
                    {"Метрика": k, "Значение": v} 
                    for k, v in analytics_data['download_stats'].items()
                ])
                download_df.to_excel(writer, sheet_name='Downloads', index=False)
            
            # Лист с дневной динамикой
            if 'daily_data' in analytics_data:
                daily_df = pd.DataFrame(analytics_data['daily_data'])
                daily_df.to_excel(writer, sheet_name='Daily Trends', index=False)
            
            # Лист с платформами
            if 'platform_stats' in analytics_data:
                platform_data = []
                for platform, stats in analytics_data['platform_stats'].items():
                    if isinstance(stats, dict):
                        row = {"Platform": platform}
                        row.update(stats)
                        platform_data.append(row)
                
                if platform_data:
                    platform_df = pd.DataFrame(platform_data)
                    platform_df.to_excel(writer, sheet_name='Platforms', index=False)
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except ImportError:
        logger.warning("pandas not available, falling back to CSV")
        return await export_analytics_to_csv(analytics_data)

def _calculate_processing_time(download) -> Optional[str]:
    """Вычислить время обработки скачивания"""
    if download.started_at and download.completed_at:
        delta = download.completed_at - download.started_at
        total_seconds = int(delta.total_seconds())
        
        if total_seconds < 60:
            return f"{total_seconds}s"
        elif total_seconds < 3600:
            minutes = total_seconds // 60
            seconds = total_seconds % 60
            return f"{minutes}m {seconds}s"
        else:
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            return f"{hours}h {minutes}m"
    
    return None

async def export_payments_to_csv(payments: List[Any]) -> StreamingResponse:
    """Экспорт платежей в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "Payment ID", "User ID", "Username", "Amount", "Currency",
        "Status", "Payment Method", "Subscription Plan", "Created At",
        "Completed At", "External ID", "Country", "Risk Score", "Is Refunded"
    ]
    writer.writerow(headers)
    
    # Данные
    for payment in payments:
        row = [
            payment.id,
            payment.payment_id,
            payment.user_id,
            getattr(payment, 'username', ''),
            float(payment.amount),
            payment.currency,
            payment.status,
            payment.payment_method,
            payment.subscription_plan or "",
            payment.created_at.isoformat() if payment.created_at else "",
            payment.completed_at.isoformat() if payment.completed_at else "",
            payment.external_payment_id or "",
            payment.country_code or "",
            payment.risk_score or "",
            "Yes" if payment.is_refunded else "No"
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def export_channels_to_csv(channels: List[Any]) -> StreamingResponse:
    """Экспорт каналов в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "Channel ID", "Channel Name", "Username", "Description",
        "Is Active", "Is Required", "Subscribers Count", "Priority",
        "Check Interval", "Created At", "Updated At"
    ]
    writer.writerow(headers)
    
    # Данные
    for channel in channels:
        row = [
            channel.id,
            channel.channel_id,
            channel.channel_name or "",
            channel.channel_username or "",
            channel.description or "",
            "Yes" if channel.is_active else "No",
            "Yes" if channel.is_required else "No",
            channel.subscribers_count or 0,
            channel.priority or 0,
            channel.check_interval_minutes or 60,
            channel.created_at.isoformat() if channel.created_at else "",
            channel.updated_at.isoformat() if channel.updated_at else ""
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=channels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def export_broadcasts_to_csv(broadcasts: List[Any]) -> StreamingResponse:
    """Экспорт рассылок в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "Title", "Status", "Target Type", "Total Recipients",
        "Sent Count", "Failed Count", "Success Rate %", "Created At",
        "Started At", "Completed At", "Created By"
    ]
    writer.writerow(headers)
    
    # Данные
    for broadcast in broadcasts:
        success_rate = 0
        if broadcast.total_recipients and broadcast.total_recipients > 0:
            success_rate = round((broadcast.sent_count or 0) / broadcast.total_recipients * 100, 2)
        
        row = [
            broadcast.id,
            broadcast.title,
            broadcast.status,
            broadcast.target_type,
            broadcast.total_recipients or 0,
            broadcast.sent_count or 0,
            broadcast.failed_count or 0,
            success_rate,
            broadcast.created_at.isoformat() if broadcast.created_at else "",
            broadcast.started_at.isoformat() if broadcast.started_at else "",
            broadcast.completed_at.isoformat() if broadcast.completed_at else "",
            getattr(broadcast, 'created_by_username', '')
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=broadcasts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

def format_file_size(size_bytes: Optional[int]) -> str:
    """Форматировать размер файла"""
    if not size_bytes:
        return ""
    
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    
    return f"{size_bytes:.1f} PB"

def format_duration(seconds: Optional[float]) -> str:
    """Форматировать длительность"""
    if not seconds:
        return ""
    
    total_seconds = int(seconds)
    
    if total_seconds < 60:
        return f"{total_seconds}s"
    elif total_seconds < 3600:
        minutes = total_seconds // 60
        secs = total_seconds % 60
        return f"{minutes}m {secs}s"
    else:
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours}h {minutes}m"

def format_percentage(value: Optional[float], decimals: int = 1) -> str:
    """Форматировать процент"""
    if value is None:
        return ""
    return f"{value:.{decimals}f}%"

async def generate_summary_report(data: Dict[str, Any]) -> str:
    """Генерация текстового отчета-сводки"""
    report_lines = []
    
    report_lines.append("=" * 60)
    report_lines.append("VIDEOBOT PRO - ADMINISTRATIVE REPORT")
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("=" * 60)
    report_lines.append("")
    
    # Общая статистика
    if 'overview' in data:
        report_lines.append("OVERVIEW:")
        for key, value in data['overview'].items():
            formatted_key = key.replace('_', ' ').title()
            report_lines.append(f"  {formatted_key}: {value}")
        report_lines.append("")
    
    # Пользователи
    if 'users' in data:
        report_lines.append("USER STATISTICS:")
        users_data = data['users']
        report_lines.append(f"  Total Users: {users_data.get('total', 0)}")
        report_lines.append(f"  Premium Users: {users_data.get('premium', 0)}")
        report_lines.append(f"  Active Users (30d): {users_data.get('active_30d', 0)}")
        report_lines.append(f"  New Users (30d): {users_data.get('new_30d', 0)}")
        report_lines.append("")
    
    # Скачивания
    if 'downloads' in data:
        report_lines.append("DOWNLOAD STATISTICS:")
        downloads_data = data['downloads']
        report_lines.append(f"  Total Downloads: {downloads_data.get('total', 0)}")
        report_lines.append(f"  Successful: {downloads_data.get('successful', 0)}")
        report_lines.append(f"  Failed: {downloads_data.get('failed', 0)}")
        success_rate = downloads_data.get('success_rate', 0)
        report_lines.append(f"  Success Rate: {success_rate}%")
        report_lines.append("")
    
    # Платежи
    if 'payments' in data:
        report_lines.append("PAYMENT STATISTICS:")
        payments_data = data['payments']
        report_lines.append(f"  Total Revenue: ${payments_data.get('total_revenue', 0):.2f}")
        report_lines.append(f"  Total Payments: {payments_data.get('total_payments', 0)}")
        report_lines.append(f"  Success Rate: {payments_data.get('success_rate', 0)}%")
        report_lines.append("")
    
    # Топ ошибки
    if 'top_errors' in data:
        report_lines.append("TOP ERRORS:")
        for i, error in enumerate(data['top_errors'][:5], 1):
            report_lines.append(f"  {i}. {error.get('message', 'Unknown')} ({error.get('count', 0)}x)")
        report_lines.append("")
    
    # Системное здоровье
    if 'system_health' in data:
        report_lines.append("SYSTEM HEALTH:")
        health = data['system_health']
        report_lines.append(f"  Overall Status: {health.get('status', 'Unknown').upper()}")
        if health.get('issues'):
            report_lines.append("  Issues:")
            for issue in health['issues']:
                report_lines.append(f"    - {issue}")
        report_lines.append("")
    
    report_lines.append("=" * 60)
    report_lines.append("End of Report")
    
    return "\n".join(report_lines)

async def export_system_logs_to_txt(logs: List[Dict[str, Any]]) -> StreamingResponse:
    """Экспорт системных логов в текстовый файл"""
    output = io.StringIO()
    
    output.write("VIDEOBOT PRO - SYSTEM LOGS\n")
    output.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    output.write("=" * 80 + "\n\n")
    
    for log_entry in logs:
        timestamp = log_entry.get('timestamp', '')
        level = log_entry.get('level', 'INFO')
        message = log_entry.get('message', '')
        module = log_entry.get('module', '')
        
        output.write(f"[{timestamp}] {level} - {module}\n")
        output.write(f"  {message}\n")
        
        if log_entry.get('details'):
            output.write(f"  Details: {log_entry['details']}\n")
        
        output.write("\n")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/plain",
        headers={
            "Content-Disposition": f"attachment; filename=system_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        }
    )